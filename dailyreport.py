from tabula import read_pdf, convert_into
import pandas as pd
import json
import os
import datetime
import glob
import openpyxl
import time

#Change working dir to EHHS/Daily Report
origin="\\\\DESKTOP-09QEAO7\\Users\\shinjuku WS-4\\Desktop\\EHHS"
os.chdir(origin)

#date
mon_list={}
today = datetime.date.today()
yesterday = today + datetime.timedelta(days=-1)
tomorrow = today + datetime.timedelta(days=1)

#Daily Status filepath
try:
    filepath = glob.glob(f"Daily Report/*/*/{yesterday:%Y%m%d}/Daily Room Status_{yesterday:%Y%m%d}.xlsx")[0]
    filepath = os.path.abspath(filepath)
except IndexError:
    print("*"*50)
    print(f"There is no Daily Status_{yesterday:%Y%m%d}.xlsx inside Daily Report folder")
    print("*"*50)

# #Daily Report filepath
# try:
#     dailyreport = glob.glob(f"Daily Report/*/*/{yesterday:%Y%m%d}/Dailyreport_{yesterday:%Y%m%d}.xlsx")[0]
#     dailyreport = os.path.abspath(dailyreport)
# except IndexError:
#     print("*"*50)
#     print(f"There is no Dailyreport_{yesterday:%Y%m%d}.xlsx inside Daily Report folder")
#     print("*"*50)

#history forcast filepath
try:
    history=glob.glob(f"History\\history_forecast_{today:%m}.pdf")[0]
except IndexError:
    print("*"*50)
    print(f"There is no history_forecast_{today:%m}.pdf inside Daily Report folder")
    print("*"*50)

#Excel cell ref
cell_for_mon=["O17","O18","O19","O20"]

data_del = [
    "P8","P9","T8","T9","K32","G33","M32","K42","M42","G43",
    "G52","D9","D19","D23","D24","D25","D26","D27","D28","D31","F31","D35","D36",
    "D39","D43","D44","D47",'H23', 'I23', 'J23', 'H24', 'I24', 'J24', 'H25', 'I25', 
    'J25', 'H26', 'I26', 'J26', 'H27', 'I27', 'J27', 'H28', 'I28', 'J28',
    "P17","P18","P19","P20",
    "Q17","Q18","Q19","Q20"
]

data_move_del ={
    "D30":"D29","F30":"F29",
}

data_need = {
    "House Use Rooms":"D19",
    "Rooms Occupied minus House Use":"P8",
    "Total In-House Persons":"D9",
    'Out of Order Rooms':"D18",
    'No Show Rooms':"D31",
    "Room Revenue":"P9",
    "today_co":"D35",
    "today_ci":"D36",
    "today_adr":"D39",
    "tmr_co":"D43",
    "tmr_ci":"D44",
    "tmr_adr":"D47",

    "1st month":"O17",
    "2nd month":"O18",
    "3rd month":"O19",
    "4th month":"O20",

    "1st month OOC":"P17",
    "1st month Rev":"Q17",
    "2nd month OOC":"P18",
    "2nd month Rev":"Q18",
    "3rd month OOC":"P19",
    "3rd month Rev":"Q19",
    "4th month OOC":"P20",
    "4th month Rev":"Q20",

    "new_1st month OOC":"T17",
    "new_1st month Rev":"U17",
    "new_2nd month OOC":"T18",
    "new_2nd month Rev":"U18",
    "new_3rd month OOC":"T19",
    "new_3rd month Rev":"U19",
    "new_4th month OOC":"T20",
    "new_4th month Rev":"U20"
}

# #cell ref for tullys
# target_cell=yesterday.day+6
# tullys_input={
#     'Out of Order Rooms':f"E{target_cell}",
#     "Rooms Occupied minus House Use":f"F{target_cell}",
#     "Total In-House Persons":f"G{target_cell}",
#     "Room Revenue":f"H{target_cell}",
#     'Other Revenue':f"I{target_cell}"
# }



#Input data from manager report and history forecast

def input_data():
    #from manager report
    try:
        mgt=glob.glob("History\\Manager*.pdf")[0]
    except IndexError:
        print(f"There is no Manager_{yesterday:%m%d}.pdf inside History folder")

    else:
        pdf_data=read_pdf(mgt,pages="all")
        manager_p1=pdf_data[0].values.tolist()
        manager_p2=pdf_data[1].values.tolist()
        manager_p1.extend(manager_p2)

        print("*"*50)
        print(f"Writing into {filepath}...")
        print("*"*50)
        # os.system("pause")

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Sheet1"]
        # tullys_wb=openpyxl.load_workbook(dailyreport)
        # sheetname=""
        # for i in tullys_wb.sheetnames:
        #     if "Input" in i:
        #         sheetname=i
        # tullys_ws=tullys_wb[sheetname]

        for i in range(4):
            try:
                temp=yesterday.replace(month=yesterday.month+i,day=1)
                mon_list[str(temp.month).zfill(2)]=f"{temp:%b}"
            except ValueError as e:
                temp=yesterday.replace(month=yesterday.month+i-12,day=1)
                mon_list[str(temp.month).zfill(2)]=f"{temp:%b}"
            finally:
                # Set the month and next 3 month
                ws[cell_for_mon[i]].value=mon_list[str(temp.month).zfill(2)]
                print(f"{ws[cell_for_mon[i]]} is set to {mon_list[str(temp.month).zfill(2)]}")
        
        # print(mon_list)

        for i in range(1,len(manager_p1)):

            try:
                if manager_p1[i][0] in data_need.keys():
                    myRange = ws[data_need[manager_p1[i][0]]]
                    myRange.value = int(manager_p1[i][1].replace(",",""))
                    print(f"Cell {data_need[manager_p1[i][0]]} is set to {myRange.value}")
                # if manager_p1[i][0] in tullys_input.keys():
                #     myRange = tullys_ws[tullys_input[manager_p1[i][0]]]
                #     myRange.value = int(manager_p1[i][1].replace(",",""))
                #     print(f"Cell {tullys_input[manager_p1[i][0]]} is set to {myRange.value}")

            except:
                pass

        #rename and get data from history forcast
        print("*"*50)
        print("Trying to get data from history forecast...")

        all_history=glob.glob(f"History\\history_forecast_*.pdf")

        if len(all_history)==0:
            print("*"*50)
            print("There is no history forcast file")
            print("*"*50)
            os.system("pause")
        else:
            for i in range(len(all_history)):
                old_path=all_history[i]
                pdf_data=read_pdf(old_path,pages="all")[0]
                pdf_data=pdf_data.values.tolist()
                file_mon=pdf_data[2][0].split()[0][0:2]
                new_path=f"History\\history_forecast_{file_mon}.pdf"
                os.rename(old_path,new_path)
                print(f"Renaming history_forecast_{file_mon}")

                if file_mon in mon_list.keys():
                    print(f"Getting data from history_forecast_{file_mon}")
                    pdf_data=read_pdf(new_path,pages="all")[0]
                    pdf_data=pdf_data.values.tolist()
                    # print(pdf_data)

                    for i in range(1,len(pdf_data)):
                        dataline=pdf_data[i]

                        # today C/I C/O ADR
                        if dataline[0].split()[0]==f"{today:%m-%d-%y}":
                            ws[data_need["today_ci"]].value=int(dataline[2])
                            ws[data_need["today_co"]].value=int(dataline[12])
                            ws[data_need["today_adr"]].value=int(dataline[11].replace(",",""))
                            print(f"Inputed today's data")
                        # tomorrow C/I C/O ADR
                        elif pdf_data[i][0].split()[0]==f"{tomorrow:%m-%d-%y}":
                            ws[data_need["tmr_ci"]].value=int(dataline[2])
                            ws[data_need["tmr_co"]].value=int(dataline[12])
                            ws[data_need["tmr_adr"]].value=int(dataline[11].replace(",",""))
                            print(f"Inputed tomorrow's data")

                        # Each month OOC and Rev now
                        elif dataline[0]=="Total":
                            if ws[data_need["1st month"]].value==mon_list[file_mon]:
                                ws[data_need["1st month OOC"]].value=int(dataline[1].replace(",",""))
                                ws[data_need["1st month Rev"]].value=int(dataline[10].replace(",",""))
                            elif ws[data_need["2nd month"]].value==mon_list[file_mon]:
                                ws[data_need["2nd month OOC"]].value=int(dataline[1].replace(",",""))
                                ws[data_need["2nd month Rev"]].value=int(dataline[10].replace(",",""))
                            elif ws[data_need["3rd month"]].value==mon_list[file_mon]:
                                ws[data_need["3rd month OOC"]].value=int(dataline[1].replace(",",""))
                                ws[data_need["3rd month Rev"]].value=int(dataline[10].replace(",",""))
                            elif ws[data_need["4th month"]].value==mon_list[file_mon]:
                                ws[data_need["4th month OOC"]].value=int(dataline[1].replace(",",""))
                                ws[data_need["4th month Rev"]].value=int(dataline[10].replace(",",""))
                            else:
                                print(f"Failed to get data from history_forecast_{file_mon}",dataline[0])

        wb.save(filepath)
        print(f"Saved to {filepath}")
        
        # tullys_ws.sheet_state = 'hidden'
        # tullys_wb.save(dailyreport)
        # try:
        #     tullys_ws=tullys_wb["E Hotel"]
        # except:
        #     tullys_ws=tullys_wb["E-Hotel"]
        # print(f"Saved to {dailyreport}")


#Create new dir and file

def make_new_dir():

    wb = openpyxl.load_workbook(filepath)
    ws = wb["Sheet1"]
    print(f"Opening {filepath} and clear data")

    for cell in data_del:
        myRange = ws[cell]
        myRange.value=None
        print(f"Cell {cell} is cleared")

    for key,value in data_move_del.items():
        ws[value]=ws[key].value
        print(f"{ws[value]} is set to {ws[key].value}")
        ws[key].value=None
        print(f"{ws[key]} is cleared")
    #set month
    new_mon_list={}
    for i in range(4):

        try:
            temp=today.replace(month=today.month+i,day=1)
            new_mon_list[str(temp.month).zfill(2)]=f"{temp:%b}"
        except ValueError:
            temp=today.replace(month=today.month+i-12,day=1)
            new_mon_list[str(temp.month).zfill(2)]=f"{temp:%b}"

        finally:
            # Set the month and next 3 month
            ws[cell_for_mon[i]].value=new_mon_list[str(temp.month).zfill(2)]
            print(f"{ws[cell_for_mon[i]]} is set to {new_mon_list[str(temp.month).zfill(2)]}")
    
    # Each month OOC and Rev now
    for new_mon in new_mon_list:
        target_pdf=glob.glob(f"History\\history_forecast_{new_mon}.pdf")
        if len(target_pdf)==1:
            pdf_data=read_pdf(target_pdf[0],pages="all")[0]
            pdf_data=pdf_data.values.tolist()
            file_mon=pdf_data[2][0].split()[0][0:2]

            for i in range(1,len(pdf_data)):
                dataline=pdf_data[i]
                if dataline[0]=="Total":
                    if ws[data_need["1st month"]].value==new_mon_list[file_mon]:
                        ws[data_need["new_1st month OOC"]].value=int(dataline[1].replace(",",""))
                        ws[data_need["new_1st month Rev"]].value=int(dataline[10].replace(",",""))
                    elif ws[data_need["2nd month"]].value==new_mon_list[file_mon]:
                        ws[data_need["new_2nd month OOC"]].value=int(dataline[1].replace(",",""))
                        ws[data_need["new_2nd month Rev"]].value=int(dataline[10].replace(",",""))
                    elif ws[data_need["3rd month"]].value==new_mon_list[file_mon]:
                        ws[data_need["new_3rd month OOC"]].value=int(dataline[1].replace(",",""))
                        ws[data_need["new_3rd month Rev"]].value=int(dataline[10].replace(",",""))
                    elif ws[data_need["4th month"]].value==new_mon_list[file_mon]:
                        ws[data_need["new_4th month OOC"]].value=int(dataline[1].replace(",",""))
                        ws[data_need["new_4th month Rev"]].value=int(dataline[10].replace(",",""))
                    else:
                        print(f"Failed to get data from history_forecast_{file_mon}")
        else:
            print(f"There is no history_forecast_{new_mon}.pdf in History folder")


    #change date
    ws["N1"].value = f"{tomorrow:%Y/%m/%d}"

    #Morning OOC and Rev
    try:
        pdf_data=read_pdf(history,pages="all")
    except NameError:
        print(f"There is no history_forecast_{today:%m}.pdf in History folder")
    else:
        hist_list=pdf_data[0].values.tolist()

        for i in range(1,len(hist_list)):
            data_want=hist_list[i]

            if data_want[0].split()[0]==f"{today:%m-%d-%y}":
                data=int(data_want[1].replace(",",""))
                ws["T8"].value=data
                print(f"Cell T8 is set to {data}")
                data=int(data_want[10].replace(",",""))
                ws["T9"].value=data
                print(f"Cell T9 is set to {data}")

    #Make new directory and file for today
    try:
        new_dir=os.path.join(origin,"Daily Report",f"{yesterday.year}年",f"{today.month}月",f"{today:%Y%m%d}")
        if os.path.isdir(new_dir)==False:
            os.makedirs(new_dir)
            print(f"New directory {new_dir} is created")

        os.chdir(new_dir)
        #save Daily Status
        save_to=os.path.join(new_dir,f"Daily Room Status_{today:%Y%m%d}.xlsx")
        wb.save(save_to)
        print(f"New file {save_to} is saved")
        
        # #Save daily report
        # if today.month==yesterday.month:
        #     wb2 = openpyxl.load_workbook(dailyreport)
        #     save_to2=os.path.join(new_dir,f"Dailyreport_{today:%Y%m%d}.xlsx")
        #     wb2.save(save_to2)
        #     print(f"New file {save_to2} is saved")
        # else:
        #     print("*"*50)
        #     print("Please find the new daily report from Mr. Hirano")
        #     print("*"*50)

    except FileExistsError as e:
        print(e)

if __name__ == "__main__":
    start=time.time()

    choice=input("""
    Please choose the function : 
    1. Get and input data from manager report and history forecast
    2. Create today's new folder and files
    >>>
    """ )
    if choice=="1":
        input_data()
    elif choice=="2":
        make_new_dir()
    else:
        choice=input("""
    Please choose the function : 
    1. Get and input data from manager report and history forecast
    2. Create today's new folder and files
    >>>
    """ )
    end=time.time()
    print(f"Finished in {round(end-start,2)} seconds")
    print("This window will be closed after 5 seconds")
    time.sleep(5)
    exit()
