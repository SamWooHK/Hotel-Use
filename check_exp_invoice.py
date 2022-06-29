import tabula
import pandas as pd
import openpyxl
from openpyxl import Workbook
import re
import os
import time


def main():
    #data from opera

    try:
        opera_file="exp.pdf"
    except:
        print("There is no exp.pdf in this folder")
    exp_list=tabula.read_pdf(opera_file,pages="all")

    opera_AR={}
    start_date=""
    end_date=""
    print(f"Fetching data from {opera_file}")
    for page in exp_list:
        content=page.values.tolist()
        for item in content:
            # print(item)
            try:
                CRS=re.sub("[^0-9]", "", item[2])
                price="".join(item[3].split(","))
                opera_AR[CRS]=int(price)
                if item[0] not in ['AR番号 :','請求先 :','日付 :','Date','日 付'] and start_date=="":
                    start_date=item[0].replace("-","/")
                end_date=content[-1][0].replace("-","/")
            except TypeError as e:
                pass
            except AttributeError as e:
                pass
            except ValueError as e:
                pass

    #data from expedia
    try:
        exp_invoice="EXP.xlsx"
    except:
        print("There is no EXP.xlsx in this folder")

    wb = openpyxl.load_workbook("EXP.xlsx")
    ws = wb["Sheet1"]


    #copy data to new workbook
    print(f"Copying data from {opera_file} to {exp_invoice}")
    new_wb = Workbook()
    new_ws = new_wb.active

    new_row=1
    new_column=1
    for row in ws.iter_rows():
        for cell in row:
            new_ws.cell(row=new_row,column=new_column).value=cell.value
            new_column+=1
        new_row+=1
        new_column=1

    #clear unwanted rows
    print("Deleting unwanted rows and data")
    for row in new_ws.rows:
        if row[0].value == "INVOICE SUMMARY":
            del_row=row[0].row

        elif row[0].value == "EXPLANATION OF REASONS:":
            bot_del=row[0].row

            new_ws.delete_rows(bot_del-1,5)
    new_ws.delete_rows(1,del_row-1)


    #style
    from openpyxl.styles.borders import Border, Side
    from openpyxl.styles.alignment import Alignment
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    title_font = Font(size=18, bold=True, underline='single')
    font = Font(size=14, bold=True, italic=False)

    fill = PatternFill(patternType='solid', fgColor='d3d3d3')

    #ref for the width
    WIDTH_dict={
        "Reservation ID":12,
        "Guest Name":12,
        "Check-In":12,
        "Check-Out":12,
        "Requested Amount":10,
    }

    start_row=[]
    end_row=[]
    total=0
    not_in_exp=[]

    print("Cross-matching data...")
    for row in new_ws.rows:

        #the start of a table
        if row[0].value == "Reservation ID":
            start_row.append(row[0].row)
            new_ws.row_dimensions[row[0].row].height = 27
            #for diff. column
            for col in range(new_ws.max_column):
                #del unwanted columns
                if row[col].value in ["Vendor Invoice Number","Confirmation Number","Payment Amount"]:
                    new_ws.delete_cols(row[col].column, 1)
                
                #set width
                elif row[col].value in WIDTH_dict.keys():
                    new_ws.column_dimensions[row[col].column_letter].width = WIDTH_dict[row[col].value]

                elif row[col].value == "Reason":
                    row[col].value = "Opera Diff"       

        #the end of a table
        elif row[0].value == "Total":
            end_row.append(row[0].row) 
            new_ws.delete_rows(row[0].row,2)     
        
        elif row[0].value in ["INVOICE SUMMARY","Invoice Number","Partner Invoice Number"]:
            row[0].font=font
            
    #get the column index of table
    table_title={}
    for row in new_ws.rows:
        if row[0].value == "Reservation ID":
            for col in range(new_ws.max_column):
                table_title[row[col].value]=row[col].column-1

    for row in new_ws.rows:
        #date
        row[table_title['Check-In']].number_format = "dd-mm-yy"
        row[table_title['Check-Out']].number_format = "dd-mm-yy"

        #total price
        if row[0].value == "Partner Invoice Number":
            #remove ,.JPY
            payment=row[table_title['Variance Amount']].value.split(".")[0]
            payment="".join(payment.split(","))
            total+=int(payment)

        #match data
        try:
            cal_price=int(row[table_title['Requested Amount']].value)//0.82
            row[table_title['Variance Amount']].value=cal_price
            row[table_title['Requested Amount']].number_format = "#,##0" 
            row[table_title['Variance Amount']].number_format = "#,##0"

            for key,value in opera_AR.items():
                if key==str(row[0].value):
                    diff=opera_AR[str(row[0].value)]-cal_price
                    if diff==0:
                        row[table_title['Opera Diff']].value="✓"
                    else:
                        row[table_title['Opera Diff']].value=diff
                        row[table_title['Opera Diff']].fill=fill
                    #remove from dict if data is matched
                    opera_AR.pop(key)

        except:
            pass

    table=zip(start_row,end_row)
        
    border = Border(top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000'),
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000')
                            )

    for area in table:
        for row_num in range(area[0],area[1]):
            for col_num in range(1,new_ws.max_column-2):  ##why max_column not changed??
                cell=new_ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal = 'center', 
                                            vertical = 'center',
                                            wrap_text = True)
                if cell.value == None:
                    cell.fill=fill

    for i in range(6):
        new_ws.insert_rows(1)
    new_ws["A1"].value="Expedia Group"
    new_ws["A1"].font=title_font
    new_ws["A2"].value=f"{start_date}-{end_date}"
    new_ws["A2"].font=title_font
    new_ws["E3"].value="合計"
    new_ws["E3"].border = border
    new_ws["E4"].value="手数料"
    new_ws["E4"].border = border
    new_ws["E5"].value="振込"
    new_ws["E5"].border = border

    new_ws["F3"].value=total//0.82
    new_ws["F3"].number_format = "#,##0" 
    new_ws["F4"].value="=F3-F5"
    new_ws["F4"].number_format = "#,##0" 
    new_ws["F5"].value=total
    new_ws["F5"].number_format = "#,##0"
    new_ws["F3"].border = border
    new_ws["F4"].border = border
    new_ws["F5"].border = border
    
    new_ws.merge_cells('F3:G3')
    new_ws.merge_cells('F4:G4')
    new_ws.merge_cells('F5:G5')


    final_row=new_ws.max_row - 2
    final_column=1

    no_data_title=new_ws.cell(row=final_row-1,column=1)
    no_data_title.value="Not in Invoice"
    no_data_title.font=title_font

    for key in opera_AR.keys():
        cell=new_ws.cell(row=final_row,column=final_column)
        cell.value=key
        cell.border = border

        if final_column>3:
            final_row+=1
            final_column=1
        else:
            final_column+=1


    print("Saving to Result.xlsx")
    new_wb.save("Result.xlsx")


if __name__ == "__main__":
    start=time.perf_counter()
    main()
    end=time.perf_counter()
    process_time=round(end-start,2)
    print(f"The program finish in {process_time} seconds")

    os.system("pause")
